from io import BytesIO
from typing import Any

from fastapi import HTTPException, UploadFile
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from sqlalchemy.orm import Session

from app.models import Category, Order, Product, ProductImage
from app.utils import calculate_discount, unique_slug


PRODUCT_COLUMNS = [
    "اسم المنتج",
    "product name",
    "القسم",
    "category",
    "السعر",
    "price",
    "الوصف",
    "description",
    "الصورة",
    "image",
    "صورة 1",
    "image1",
    "صورة 2",
    "image2",
    "صورة 3",
    "image3",
    "صورة 4",
    "image4",
    "صورة 5",
    "image5",
    "السعر بعد الخصم",
    "discount price",
    "نسبة الخصم",
    "discount percentage",
    "متوفر",
    "available",
    "مميز",
    "featured",
]

HEADER_ALIASES = {
    "اسم المنتج": "name_ar",
    "product name": "name_ar",
    "name": "name_ar",
    "القسم": "category",
    "category": "category",
    "السعر": "price",
    "price": "price",
    "الوصف": "description_ar",
    "description": "description_ar",
    "الصورة": "image",
    "image": "image",
    "السعر بعد الخصم": "discount_price",
    "discount price": "discount_price",
    "نسبة الخصم": "discount_percentage",
    "discount percentage": "discount_percentage",
    "متوفر": "is_available",
    "available": "is_available",
    "مميز": "is_featured",
    "featured": "is_featured",
}


def as_bool(value: Any, default: bool = False) -> bool:
    if value is None or value == "":
        return default
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() in {"yes", "true", "1", "y", "نعم", "متوفر", "مميز"}


def as_float(value: Any) -> float | None:
    if value in (None, ""):
        return None
    return float(str(value).replace("JD", "").replace("د.أ", "").strip())


def normalize_headers(headers: list[Any]) -> dict[int, str]:
    normalized: dict[int, str] = {}
    for index, header in enumerate(headers):
        key = str(header or "").strip().lower()
        if not key:
            continue
        if key.startswith("صورة") or key.startswith("image"):
            normalized[index] = "gallery_image"
        else:
            normalized[index] = HEADER_ALIASES.get(key, key)
    return normalized


def parse_product_workbook(upload: UploadFile) -> tuple[list[dict[str, Any]], list[str]]:
    try:
        workbook = load_workbook(upload.file, data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Could not read Excel file: {exc}") from exc
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return [], ["Excel file is empty"]
    header_map = normalize_headers(list(rows[0]))
    parsed: list[dict[str, Any]] = []
    errors: list[str] = []
    for row_number, row in enumerate(rows[1:], start=2):
        if not any(row):
            continue
        item: dict[str, Any] = {"gallery_images": []}
        for index, value in enumerate(row):
            field = header_map.get(index)
            if not field:
                continue
            if field == "gallery_image":
                if value:
                    item["gallery_images"].append(str(value).strip())
            else:
                item[field] = value
        if not item.get("name_ar"):
            errors.append(f"Row {row_number}: missing product name")
            continue
        if not item.get("category"):
            errors.append(f"Row {row_number}: missing category")
            continue
        try:
            item["price"] = as_float(item.get("price"))
        except ValueError:
            errors.append(f"Row {row_number}: invalid price")
            continue
        if not item.get("price"):
            errors.append(f"Row {row_number}: price is required")
            continue
        item["discount_price"] = as_float(item.get("discount_price"))
        item["discount_percentage"] = int(item.get("discount_percentage") or 0)
        item["is_available"] = as_bool(item.get("is_available"), True)
        item["is_featured"] = as_bool(item.get("is_featured"), False)
        item["is_offer"] = bool(item["discount_price"] or item["discount_percentage"])
        image = item.get("image")
        if image:
            item["gallery_images"].insert(0, str(image).strip())
        parsed.append(item)
    return parsed, errors


def import_products(db: Session, upload: UploadFile, commit: bool = False) -> dict[str, Any]:
    rows, errors = parse_product_workbook(upload)
    if errors or not commit:
        return {"ok": not errors, "created": 0, "updated": 0, "rows": rows[:100], "errors": errors}

    created = 0
    updated = 0
    for row in rows:
        category_name = str(row["category"]).strip()
        category = db.query(Category).filter(Category.name_ar == category_name).first()
        if not category:
            category = Category(
                name_ar=category_name,
                name_en=category_name,
                slug=unique_slug(db, Category, category_name),
                description=f"قسم {category_name} في غالب مول",
                is_active=True,
            )
            db.add(category)
            db.flush()
        product = db.query(Product).filter(Product.name_ar == str(row["name_ar"]).strip()).first()
        if product:
            updated += 1
        else:
            created += 1
            product = Product(name_ar=str(row["name_ar"]).strip(), slug=unique_slug(db, Product, str(row["name_ar"])))
            db.add(product)
        product.category_id = category.id
        product.price = float(row["price"])
        product.discount_price = row.get("discount_price")
        product.discount_percentage = row.get("discount_percentage") or calculate_discount(product.price, product.discount_price)
        product.description_ar = str(row.get("description_ar") or "")
        product.main_image = row["gallery_images"][0] if row.get("gallery_images") else None
        product.is_available = bool(row.get("is_available", True))
        product.is_featured = bool(row.get("is_featured", False))
        product.is_offer = bool(row.get("is_offer", False))
        product.images.clear()
        for index, image_url in enumerate(row.get("gallery_images", [])):
            product.images.append(ProductImage(image_url=image_url, sort_order=index, alt_text=product.name_ar))
    db.commit()
    return {"ok": True, "created": created, "updated": updated, "rows": rows[:100], "errors": []}


def create_product_template() -> BytesIO:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Products"
    headers = [
        "اسم المنتج",
        "القسم",
        "السعر",
        "الوصف",
        "الصورة",
        "صورة 1",
        "صورة 2",
        "صورة 3",
        "السعر بعد الخصم",
        "نسبة الخصم",
        "متوفر",
        "مميز",
    ]
    sheet.append(headers)
    sheet.append([
        "بلوزة ستاتي ناعمة",
        "ستاتي",
        9.99,
        "بلوزة يومية بخامة مريحة وموديل عصري.",
        "https://example.com/main.jpg",
        "https://example.com/angle-1.jpg",
        "",
        "",
        7.99,
        "",
        "نعم",
        "نعم",
    ])
    for cell in sheet[1]:
        cell.fill = PatternFill("solid", fgColor="E85D25")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    sheet.freeze_panes = "A2"
    for column in sheet.columns:
        width = max(len(str(cell.value or "")) for cell in column) + 4
        sheet.column_dimensions[column[0].column_letter].width = min(width, 28)
    table = Table(displayName="ProductsImportTemplate", ref=f"A1:L{sheet.max_row}")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium4", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    sheet.add_table(table)
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def export_products(db: Session) -> BytesIO:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Products"
    headers = [
        "id",
        "اسم المنتج",
        "English name",
        "القسم",
        "السعر",
        "السعر بعد الخصم",
        "نسبة الخصم",
        "الوصف",
        "الصورة",
        "متوفر",
        "مميز",
        "عرض",
    ]
    sheet.append(headers)
    products = db.query(Product).order_by(Product.created_at.desc()).all()
    for product in products:
        sheet.append([
            product.id,
            product.name_ar,
            product.name_en,
            product.category.name_ar if product.category else "",
            product.price,
            product.discount_price,
            product.discount_percentage,
            product.description_ar,
            product.main_image,
            "نعم" if product.is_available else "لا",
            "نعم" if product.is_featured else "لا",
            "نعم" if product.is_offer else "لا",
        ])
    for cell in sheet[1]:
        cell.fill = PatternFill("solid", fgColor="1F2937")
        cell.font = Font(color="FFFFFF", bold=True)
    sheet.freeze_panes = "A2"
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def export_orders(db: Session) -> BytesIO:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Orders"
    headers = ["Order number", "Customer", "Phone", "Address", "Status", "Total JD", "Created", "Items"]
    sheet.append(headers)
    orders = db.query(Order).order_by(Order.created_at.desc()).all()
    for order in orders:
        items = " | ".join(f"{item.product_name_snapshot} x{item.quantity}" for item in order.items)
        sheet.append([
            order.order_number,
            order.customer_name,
            order.customer_phone,
            order.customer_address,
            order.status,
            order.total_amount,
            order.created_at.isoformat(timespec="minutes"),
            items,
        ])
    for cell in sheet[1]:
        cell.fill = PatternFill("solid", fgColor="E85D25")
        cell.font = Font(color="FFFFFF", bold=True)
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output

